import json
from openpyxl import load_workbook

file_path = "C:\\Users\\sawan\\Downloads\\TY Div C.xlsx"

try:
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    data = []
    for row in sheet.iter_rows(values_only=True):
        row_data = [str(cell).strip() if cell else "" for cell in row]
        if any(row_data):
            data.append(row_data)

    for i in range(min(15, len(data))):
        print(data[i])

    with open("C:\\Users\\sawan\\Downloads\\ty_div_c.json", "w") as f:
        json.dump(data, f)
except Exception as e:
    print(f"Error: {e}")
